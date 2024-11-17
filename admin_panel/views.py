import io
import re
from django.shortcuts import render, redirect,get_object_or_404,HttpResponse
from django.contrib.auth.models import User
from django.contrib.auth import authenticate,login,logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from ekart.models import Product,EndUser,Order,OrderItem,Category,Payment
from .models import Supplier,Slide
from django.db.models import Sum
from .forms import ProductForm,CategoryForm,SupplierForm,SlideForm
from xhtml2pdf import pisa
from django.template.loader import get_template
import json
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import xml.etree.ElementTree as ET
import openpyxl


def admin_panel(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)

        if user is not None and user.is_staff: 
            login(request, user)
            return redirect('admin_dashboard') 
        else:
            messages.error(request, 'Invalid username or password')
    
    return render(request, 'adminpanel.html')

@login_required
def admin_dashboard(request):
    total_sales = Order.objects.filter(status='Delivered').aggregate(Sum('total_price'))['total_price__sum'] or 0
    total_sales = float(total_sales)  

    total_stock = Product.objects.aggregate(Sum('stock_quantity'))['stock_quantity__sum'] or 0
    total_stock = int(total_stock) 

    new_orders_count = Order.objects.filter(status__in=['Pending', 'Processing']).count()

    sales_data = (
        Order.objects.filter(status='Delivered')
        .values('order_date__month')
        .annotate(total_sales=Sum('total_price'))
        .order_by('order_date__month')
    )

    months = [str(item['order_date__month']) for item in sales_data]
    sales_values = [float(item['total_sales']) for item in sales_data]

    
    products = Product.objects.all()
    product_names = [product.title for product in products]
    stock_quantities = [product.stock_quantity for product in products]

    context = {
        'total_sales': total_sales,
        'total_stock': total_stock,
        'new_orders_count': new_orders_count,
        'months': json.dumps(months),
        'sales_values': json.dumps(sales_values),
        'product_names': json.dumps(product_names),
        'stock_quantities': json.dumps(stock_quantities),
    }

    return render(request, 'admin_dashboard.html', context)
@login_required
def admin_products(request):
    products = Product.objects.all()
    if request.method == 'POST' and 'add_product' in request.POST:
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Product added successfully!')
            return redirect('admin_products')
    else:
        form = ProductForm()

    return render(request, 'admin_products.html', {
        'products': products,
        'form': form,
    })

@login_required
def edit_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, 'Product updated successfully!')
            return redirect('admin_products')
    else:
        form = ProductForm(instance=product)

    return render(request, 'edit_product.html', {'form': form, 'product': product})


@login_required
def delete_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    if request.method == 'POST':
        product.delete()
        messages.success(request, 'Product deleted successfully!')
        return redirect('admin_products')
    return render(request, 'delete_confirmation.html', {'product': product})
@login_required
def sales(request):
    orders = Order.objects.all().prefetch_related('items')

    sales_data = []
    for order in orders:
        for item in order.items.all():
            sales_data.append({
                'order': order,
                'product': item.product,
                'quantity': item.quantity,
                'total_price': item.quantity * item.price,
            })

    context = {
        'sales': sales_data,
    }

    return render(request, 'admin_sales.html', context)

@login_required
def update_order_status(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    
    if request.method == 'POST':
        status = request.POST.get('status')
        if status:
            order.status = status
            order.save()
    
    return redirect('sales') 

@login_required
def update_delivery_date(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    
    if request.method == 'POST':
        estimated_delivery_date = request.POST.get('estimated_delivery_date')
        if estimated_delivery_date:
            order.estimated_delivery_date = estimated_delivery_date
            order.save()
    
    return redirect('sales')
@login_required
def print_receipt(request, order_id):
    order = get_object_or_404(Order, order_id=order_id)
    
    context = {
        'order': order,
    }
    return render(request, 'print_receipt.html', context)


@login_required
def admin_customers(request):
    customers = EndUser.objects.all()  
    return render(request, 'admin_customers.html', {'customers': customers})


def logout_view(request):
    logout(request)
    return redirect('adminpanel')  

def download_receipt_pdf(request, order_id):
   
    order = get_object_or_404(Order, order_id=order_id)
    
    template = get_template("print_receipt.html")
    html = template.render({"order": order})
    
    response = HttpResponse(content_type="application/pdf")
    response['Content-Disposition'] = f'attachment; filename="Order_{order_id}_Receipt.pdf"'
    
    pisa_status = pisa.CreatePDF(html, dest=response)
    
    if pisa_status.err:
        return render(request, "print_receipt.html", {"order": order, "error_message": "There was an error generating the PDF."})
    
    return response

def category_list(request):
    categories = Category.objects.all()
    return render(request, 'category_list.html', {'categories': categories})

def add_category(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('category_list')
    else:
        form = CategoryForm()
    return render(request, 'category_form.html', {'form': form, 'title': 'Add Category'})

def edit_category(request, category_id):
    category = get_object_or_404(Category, id=category_id)
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            return redirect('category_list')
    else:
        form = CategoryForm(instance=category)
    return render(request, 'category_form.html', {'form': form, 'title': 'Edit Category'})

def delete_category(request, category_id):
    category = get_object_or_404(Category, id=category_id)
    if request.method == 'POST':
        category.delete()
        return redirect('category_list')
    return render(request, 'category_confirm_delete.html', {'category': category})

def supplier_list(request):
    suppliers = Supplier.objects.all()
    return render(request, 'supplier_list.html', {'suppliers': suppliers})

def add_supplier(request):
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('supplier_list')
    else:
        form = SupplierForm()
    return render(request, 'supplier_form.html', {'form': form, 'title': 'Add Supplier'})

def edit_supplier(request, supplier_id):
    supplier = get_object_or_404(Supplier, id=supplier_id)
    if request.method == 'POST':
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            return redirect('supplier_list')
    else:
        form = SupplierForm(instance=supplier)
    return render(request, 'supplier_form.html', {'form': form, 'title': 'Edit Supplier'})

def delete_supplier(request, supplier_id):
    supplier = get_object_or_404(Supplier, id=supplier_id)
    if request.method == 'POST':
        supplier.delete()
        return redirect('supplier_list')
    return render(request, 'supplier_confirm_delete.html', {'supplier': supplier})

def payment_details(request):
    payments = Payment.objects.all()
    return render(request, 'payment_details.html', {'payments': payments})


def export_payments_pdf(request):
    buffer = io.BytesIO()
    pdf_canvas = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    pdf_canvas.drawString(100, height - 50, "Payment Statement")
    
    y_position = height - 100
    payments = Payment.objects.all()

    for payment in payments:
        pdf_canvas.drawString(50, y_position, f"User: {payment.user}")
        pdf_canvas.drawString(200, y_position, f"Order ID: {payment.order_id}")
        pdf_canvas.drawString(350, y_position, f"Amount: ₹{payment.amount}")
        pdf_canvas.drawString(450, y_position, f"Status: {payment.status}")
        y_position -= 20
        if y_position < 50: 
            pdf_canvas.showPage()
            y_position = height - 50

    pdf_canvas.save()
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="payment_statement.pdf"'
    return response

def export_payments_excel(request):
    payments = Payment.objects.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payments"
    headers = [
        "User", "OrderID", "Amount", "PaymentMethod", "TransactionID", "Status", "Date"
    ]

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    for row_num, payment in enumerate(payments, 2):
        ws.cell(row=row_num, column=1, value=payment.user.email)
        ws.cell(row=row_num, column=2, value=payment.order_id)
        ws.cell(row=row_num, column=3, value=str(payment.amount))
        ws.cell(row=row_num, column=4, value=payment.get_payment_method_display())
        ws.cell(row=row_num, column=5, value=payment.transaction_id or "N/A")
        ws.cell(row=row_num, column=6, value=payment.status)
        ws.cell(row=row_num, column=7, value=payment.created_at.strftime("%Y-%m-%d %H:%M:%S"))
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="payment_statement.xlsx"'
    wb.save(response)
    
    return response
def advertisement_management(request):
    slides = Slide.objects.all()  # Retrieve all slides
    form = SlideForm()  # Default form initialization for GET requests and any fallback cases

    if request.method == 'POST':
        # Adding a new slide
        if 'add_slide' in request.POST:
            form = SlideForm(request.POST, request.FILES)
            if form.is_valid():
                form.save()
                messages.success(request, "Slide added successfully!")
                return redirect('advertisement_management')  # Redirect after adding to clear form data
            else:
                messages.error(request, f"Error adding slide: {form.errors}")  # Show form errors in messages

        # Editing an existing slide
        elif 'edit_slide' in request.POST:
            slide_id = request.POST.get('slide_id')
            slide = get_object_or_404(Slide, id=slide_id)
            form = SlideForm(request.POST, request.FILES, instance=slide)
            if form.is_valid():
                form.save()
                messages.success(request, "Slide updated successfully!")
                return redirect('advertisement_management')
            else:
                messages.error(request, f"Error updating slide: {form.errors}")

        # Deleting a slide
        elif 'delete_slide' in request.POST:
            slide_id = request.POST.get('slide_id')
            slide = get_object_or_404(Slide, id=slide_id)
            slide.delete()
            messages.success(request, "Slide deleted successfully!")
            return redirect('advertisement_management')

    # Render the template with the slides and form
    return render(request, 'manage_advertisements.html', {
        'slides': slides,
        'form': form,
    })
